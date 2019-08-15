#!pip install openpyxl 
#!pip install pandas

from openpyxl import load_workbook 
import pandas as pd
import numpy as np 
import utils
import re

FIRST_MEETUP_COLUMN_ID = 7
FIRST_MEMBER_ROW_ID = 8

MEETUP_NAME_ROW_ID = 3

MEMBER_ORDEM_COLUMN_ID = 1
MEMBER_NAME_COLUMN_ID = 2
MEMBER_CONTACT_URL_COLUMN_ID = 4

class PresenceSpreadsheet(object):

    def __init__(self, filename):
        # self.wb = load_workbook('./Lista de Presenças.xlsx') 
        # sheet_name = 'Presenças'
        self.filename = filename

    def load(self) :
        '''
            Load in the workbook and sheet
        ''' 
        self.wb = load_workbook(self.filename) 
        self.sheet = self.wb.get_active_sheet()
        
    def read_all(self):
        self.members = self.read_members()
        self.meetups = self.read_meetups()
        self.presences = self.read_presences()
    
    def read_members(self, start_at = FIRST_MEMBER_ROW_ID):
        '''
        Reads all members rows
        '''
        self.members = []
        
        # get first member name
        row_id = start_at
        ordem_cell = self.sheet.cell(row= row_id, column = MEMBER_ORDEM_COLUMN_ID)
        name_cell  = self.sheet.cell(row= row_id, column = MEMBER_NAME_COLUMN_ID)

        # iterate over all members
        while utils.isNotBlank(name_cell.value):

            ordem_cell = self.sheet.cell(row= row_id, column = MEMBER_ORDEM_COLUMN_ID)
            name_cell  = self.sheet.cell(row = row_id, column = MEMBER_NAME_COLUMN_ID) 
            
            # check value
            if(name_cell.value):

                contact_url_cell = self.sheet.cell( row = row_id, 
                                                    column = MEMBER_CONTACT_URL_COLUMN_ID)
                
                # meetup_user_id = utils.parse_url_object_id( contact_url_cell.value )
                meetup_user_id = re.search(r'members/(.*?)/profile', contact_url_cell.value).group(1)
                
                self.members.append({
                                    "uid": row_id, 
                                    "ordem":ordem_cell.value,
                                    "name": name_cell.value,
                                    "contact_url": contact_url_cell.value, 
                                    "meetup_user_id": meetup_user_id
                                    }) 
            
            # next row id
            row_id += 1

        return self.members


    def read_meetups(self, start_at = FIRST_MEETUP_COLUMN_ID):
        '''
        Reads all members rows
        '''    
        self.meetups = []
        
        # get first meetup name
        col_id = start_at
        meetup_edition_cell = self.sheet.cell( row = MEETUP_NAME_ROW_ID , 
                                            column = col_id)

        while utils.isNotBlank(meetup_edition_cell.value):
            meetup_edition_cell = self.sheet.cell( row = MEETUP_NAME_ROW_ID, 
                                                column = col_id)

            if(meetup_edition_cell.value):
                self.meetups.append({"uid": col_id, 
                                     "meetup_edition": meetup_edition_cell.value}) 
                
            col_id += 1

        return self.meetups    

    def persist_presences(self, presences):
        pass
    
    def read_presences(self):
            
        self.presences = np.zeros((len(self.members), len(self.meetups)), dtype=int)
        
        for member in self.members:
            for meetup in self.meetups:

                presence_cell = self.sheet.cell(row= int(member["uid"]) , 
                                                column = int(meetup["uid"]))        

                if(utils.isNotBlank(presence_cell.value)):
                    if(utils.isPresentSymbol(presence_cell.value)):
                        self.presences[int(member["uid"]) - FIRST_MEMBER_ROW_ID, 
                                       int(meetup["uid"]) - FIRST_MEETUP_COLUMN_ID ] = 1

        return self.presences